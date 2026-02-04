import argparse
import datetime as dt
import pyodbc
from openpyxl import load_workbook
import pyperclip



SQL_SERVER=r"USHYDHEMHEMA01\SQLEXPRESS"
SQL_DATABASE="POC"

TEMPLATE_XLSX="run_template.xlsx"
OUTPUT_XLSX="output.xlsx"

#MODEL_SHEETS=["FA STAT","VA STAT","FIA STAT","VA MRB","JAPAN MRB"]

CELL_MODEL_NAME="B1"
CELL_FOLDER="B2"
CELL_FARM_NAME="B3"
CELL_FARM_TYPE="B4"
CELL_CORES="B5"
CELL_ITERATION_NO="B6"
CELL_VAL_DATE="B7"
CELL_FINAL_COMMAND="B9"

DUMMY_VALUES={
    CELL_FOLDER: r"C:\Users\Datasets\Model\{model}" ,
    CELL_FARM_NAME: "STANDARD",
    CELL_FARM_TYPE: "ELINK",
    CELL_CORES: 120,
}

PEGA_COMMAND_PREFIX="RUN_MODEL"

def connect_sql_widows_auth():
    conn_str=(
        "DRIVER={ODBC Driver 17 for SQL Server};"
        f"SERVER={SQL_SERVER};"
        f"DATABASE={SQL_DATABASE};"
        "Trusted_Connection=yes;"
        "Encrypt=no;"
        "TrustedServerCertificate=yes;"
    )
    return pyodbc.connect(conn_str)

def get_latest_run_id(conn,model_name: str,as_of:dt.datetime | None=None) -> str:
    cur=conn.cursor()
    if as_of is None:
        cur.execute(
           """ 
            SELECT TOP 1 RUNID
            FROM dbo.ModelRuns
            where ModelName=?
            ORDER BY RunDate DESC, CreatedAt DESC
            """,model_name,
        )
    else:
       cur.execute(
            """
            SELECT TOP 1 RUNID
            FROM dbo.ModelRuns
            where ModelName=?
            AND RunDate<=?
            ORDER BY RunDate DESC, CreatedAt DESC
            """,model_name,as_of,
        ) 
    row=cur.fetchone()
    if not row:
        raise ValueError(f"No RunId found in DB for ModelName='{model_name}'")
    return str(row[0])

def valuation_date_mmyyyy(today=None)->str:
    today=today or dt.date.today()
    return today.strftime("%m%Y")

def build_final_command(model:str,folder:str,farm_name:str,farm_type:str,
                        cores,iteration_id:str,val_date:str)->str:
    parts=[
        PEGA_COMMAND_PREFIX,
        f"MODEL={model}",
        f"FOLDER={folder}",
        f"FARM_NAME={farm_name}",
        f"FARM_TYPE={farm_type}",
        f"CORES={cores}",
        f"ITERATION_ID={iteration_id}",
        f"VALUATION_DATE={val_date}",
    ]
    return ",".join(parts)

def fill_sheet(ws,model_name:str,iteration_id:str,val_date:str) -> str:
    ws[CELL_MODEL_NAME]=model_name

    for cell,value in DUMMY_VALUES.items():
        if isinstance(value,str):
            ws[cell]=value.format(model=model_name.replace(" ","_"))
        else:
            ws[cell]=value
    
    ws[CELL_ITERATION_NO]=iteration_id
    ws[CELL_VAL_DATE]=val_date

    folder=str(ws[CELL_FOLDER].value)
    farm_name=str(ws[CELL_FARM_NAME].value)
    farm_type=str(ws[CELL_FARM_TYPE].value)
    cores=ws[CELL_CORES].value

    final_cmd=build_final_command(
        model=model_name,
        folder=folder,
        farm_name=farm_name,
        farm_type=farm_type,
        cores=cores,
        iteration_id=iteration_id,
        val_date=val_date,
    )
    ws[CELL_FINAL_COMMAND]=final_cmd
    return final_cmd

def parse_args():
    p=argparse.ArgumentParser(description="POC: DB-> Excel -> comma-seperated Pega Command(single model).")
    p.add_argument("model",help='Model sheet name,e.g "FA STAT"')
    p.add_argument("--asof",help='optional as-of datetim, eg"2026-02-03 12:00:00"',default=None)
    return p.parse_args()

def main():
    args=parse_args()
    model=args.model.strip()

    as_of_dt=None
    if args.asof:
        as_of_dt=dt.datetime.strptime(args.asof,"%Y-%m-%d %H:%M:%S")
    val_date=valuation_date_mmyyyy()

    conn=connect_sql_widows_auth()
    run_id=get_latest_run_id(conn,model,as_of=as_of_dt)

    wb=load_workbook(TEMPLATE_XLSX)
    if model not in wb.sheetnames:
        raise ValueError(f"Sheet '{model}' not found.Available:{wb.sheetnames}")
    
    ws=wb[model]
    final_cmd=fill_sheet(ws,model,run_id,val_date)

    wb.save(OUTPUT_XLSX)
    pyperclip.copy(final_cmd)

    print("POC complete (single model).")
    print(f"-Model: {model}")
    print(f"- RunID (Iteration Id): {run_id}")
    print(f"- Valuation Date(MMYYYY): {val_date}")
    print(f"- Saved workbook: {OUTPUT_XLSX}")
    print("- Final command(also copied to clipboard):")
    print(final_cmd)

if __name__=="__main__":
    main()